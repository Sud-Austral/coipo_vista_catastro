"""Convierte el libro de homologacion a CSV versionados, uno por hoja.

POR QUE NO SE LEE EL .XLSX DIRECTAMENTE DESDE EL ETL. Un insumo binario que
nadie puede diffear es exactamente la trampa que este repo evita en todo lo
demas: si manana alguien cambia una celda, `git diff` no ensena nada y la cifra
publicada cambia sin rastro. Con los CSV versionados, cada correccion de la
Unidad se ve como lo que es -- una linea que cambia-- y queda en el historial.
Ademas el .xlsx vive en INSUMO/, que no esta versionado.

Se conserva TODO, incluidas las filas `sin_cambio`: la tabla no es solo un mapa
de correcciones, es el CATALOGO CERRADO de valores admitidos. `build_bin.py`
revienta si el origen trae un valor que no este aqui, y eso solo funciona si las
filas sin cambio tambien estan.

CUIDADO AL LEER ESTOS CSV. El codigo de especie del rauli es literalmente «NA».
pandas y R lo convierten en nulo por omision y lo pierden -- es el unico caso en
todo el vocabulario, y son 3.654 poligonos con 89.994,21 ha de una especie
comercial. Se lee con el modulo `csv`, que no interpreta nada, o con
`keep_default_na=False`. Por lo mismo el codigo de especie NO se pliega a
mayusculas en ningun punto: «AB» es Abies, «Ab» es Adesmia boronioides y «ab» es
Calceolaria biflora, y un upper() descuidado fusionaria 206 especies distintas.

Uso:  python ETL/homologacion/desde_xlsx.py [ruta al .xlsx]
"""

import csv
import os
import sys

import openpyxl

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(os.path.dirname(AQUI))
LIBRO = os.path.join(RAIZ, "INSUMO", "homologacion_catastro_1.xlsx")

sys.stdout.reconfigure(encoding="utf-8")

# La primera celda de la fila de cabecera. Las hojas llevan dos o tres lineas de
# titulo y glosa antes, y esas lineas SE PIERDEN a proposito: son prosa para el
# lector del libro, y su sitio en el repo es este docstring y los comentarios del
# ETL, no una fila fantasma dentro del CSV.
CABECERAS = ("valor_origen", "cod_origen", "dimension", "especie_cod", "genero")


def exportar(wb, hoja):
    ws = wb[hoja]
    filas = [list(r) for r in ws.iter_rows(values_only=True)]
    inicio = next((i for i, r in enumerate(filas)
                   if r and str(r[0]).strip() in CABECERAS), None)
    if inicio is None:
        return None
    cab = [str(c).strip() for c in filas[inicio] if c is not None]
    ancho = len(cab)
    cuerpo = []
    for r in filas[inicio + 1:]:
        if not r or r[0] is None or str(r[0]).strip() == "":
            continue
        cuerpo.append(["" if c is None else str(c).strip() for c in r[:ancho]])

    destino = os.path.join(AQUI, f"{hoja}.csv")
    with open(destino, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(cab)
        w.writerows(cuerpo)
    return len(cuerpo), ancho


def main():
    libro = sys.argv[1] if len(sys.argv) > 1 else LIBRO
    if not os.path.isfile(libro):
        sys.exit(f"no existe el libro: {libro}")
    wb = openpyxl.load_workbook(libro, read_only=True, data_only=True)

    total = 0
    for hoja in wb.sheetnames:
        r = exportar(wb, hoja)
        if r is None:
            print(f"  --     {hoja:<26} sin fila de cabecera reconocible, se omite")
            continue
        filas, cols = r
        total += filas
        print(f"  ok     {hoja:<26} {filas:>4} filas x {cols} columnas")
    print(f"\n{total} filas exportadas a {os.path.relpath(AQUI, RAIZ)}")


if __name__ == "__main__":
    main()
